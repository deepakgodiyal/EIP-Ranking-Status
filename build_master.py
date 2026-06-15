import openpyxl, json, os
from collections import defaultdict, Counter

UP = '/sessions/inspiring-stoic-davinci/mnt/uploads'
OLD = UP+'/1 April Ranking Sheet-1628a2e5.xlsx'
JUN11 = UP+'/11-June-2026 - EMR Ranking-001a2ab3.xlsx'
PEOPLE = {'Harsh','Dubey','Gagan','Saurabh'}

# Master sheet's column headers are WRONG. User-confirmed real dates for cols 18..26:
SEED_LABELS = ['01 Apr','07 Apr','13 Apr','23 Apr','05 May','13 May','18 May','05 Jun','09 Jun']

def seed_master():
    ws = openpyxl.load_workbook(OLD, data_only=True)['Keywords Ranking']
    cols = list(range(18,27))  # 9 columns = the 9 SEED_LABELS
    assert len(cols)==len(SEED_LABELS)
    def parse(v):
        if v is None: return None
        if isinstance(v,(int,float)): return int(v)
        s=str(v).strip()
        if s in ('','#N/A'): return None
        if s in ('-','New'): return '-'
        try: return int(float(s))
        except: return '-'
    kw={}
    for r in range(2, ws.max_row+1):
        k=ws.cell(r,2).value; g=ws.cell(r,4).value
        if not k or not g: continue
        k=str(k).strip()
        kw[k]={'kw':k,'grp':str(g).strip(),'r':[parse(ws.cell(r,c).value) for c in cols],
               'searches':None,'url':None}
    return {'dates':list(SEED_LABELS),'kw':kw}

def add_datefile(master, path, label):
    ws = openpyxl.load_workbook(path, data_only=True)['Sheet1']
    hdr={ws.cell(1,c).value:c for c in range(1, ws.max_column+1)}
    cK,cS,cR,cU,cT=hdr['Keyword'],hdr['# of Searches'],hdr['Rank'],hdr['URL Found'],hdr['Tags']
    def prank(v):
        if v is None: return None
        if isinstance(v,(int,float)): return int(v)
        s=str(v).strip()
        if s.lower().startswith('not in'): return '-'
        if s in ('','-'): return '-'
        try: return int(float(s))
        except: return '-'
    def psrch(v):
        if v is None or v=='N/A' or v=='': return None
        try: return int(float(v))
        except: return None
    if label not in master['dates']:
        master['dates'].append(label)
        for k in master['kw'].values(): k['r'].append(None)
    di=master['dates'].index(label)
    for r in range(2, ws.max_row+1):
        k=ws.cell(r,cK).value
        if not k: continue
        k=str(k).strip()
        grp=str(ws.cell(r,cT).value).strip() if ws.cell(r,cT).value else None
        if k not in master['kw']:
            master['kw'][k]={'kw':k,'grp':grp,'r':[None]*len(master['dates']),'searches':None,'url':None}
        rec=master['kw'][k]
        while len(rec['r'])<len(master['dates']): rec['r'].append(None)
        rec['r'][di]=prank(ws.cell(r,cR).value)
        if grp: rec['grp']=grp
        rec['searches']=psrch(ws.cell(r,cS).value)
        u=ws.cell(r,cU).value
        if u: rec['url']=u   # keep existing URL if a future date file has none
    return master

master=seed_master()
add_datefile(master, JUN11, '11 Jun')
json.dump(master, open('ranking_data.json','w'), default=str)
print('dates:', master['dates'])
print('keywords:', len(master['kw']))

# derive dashboard dataset
dates=master['dates']; groups={}
def nb(): return {'b3':0,'b4_10':0,'b11_30':0,'b_oth':0,'total':0,'live':0,'notlive':0,'newlive':0}
kw_rows=[]
for rec in master['kw'].values():
    grp=rec['grp']
    if grp not in groups: groups[grp]={'isPerson':grp in PEOPLE,'series':[nb() for _ in dates]}
    series=groups[grp]['series']; prev_live=None
    for i,v in enumerate(rec['r']):
        s=series[i]; s['total']+=1
        if v is None: s['notlive']+=1
        else:
            s['live']+=1
            if v=='-': s['b_oth']+=1
            elif isinstance(v,int) and v>30: s['b_oth']+=1
            elif v<=3: s['b3']+=1
            elif v<=10: s['b4_10']+=1
            else: s['b11_30']+=1
            if prev_live is False: s['newlive']+=1
        prev_live=(v is not None)
    def num(x): return x if isinstance(x,int) else None
    lst=num(rec['r'][-1]); prv=num(rec['r'][-2]) if len(rec['r'])>=2 else None
    lr=rec['r'][-1]; pr=rec['r'][-2] if len(rec['r'])>=2 else None
    if lr is None: st='notlive'
    elif pr is None and lr is not None: st='new'
    elif lst is not None and prv is not None: st='up' if lst<prv else ('down' if lst>prv else 'same')
    elif lst is not None and prv is None: st='up'
    elif lst is None and prv is not None: st='down'
    else: st='same'
    chg=(prv-lst) if (lst is not None and prv is not None) else None
    kw_rows.append({'kw':rec['kw'],'grp':grp,'r':rec['r'],'searches':rec['searches'],'url':rec['url'],'chg':chg,'st':st})

out={'dates':dates,'groups':groups,'people':sorted(PEOPLE),'kw':kw_rows}
json.dump(out, open('data4.json','w'), default=str)
print('size',os.path.getsize('data4.json'),'status',Counter(k['st'] for k in kw_rows))
for p in sorted(PEOPLE):
    s=groups[p]['series']
    print(p,'09Jun top3',s[-2]['b3'],'| 11Jun top3',s[-1]['b3'])
